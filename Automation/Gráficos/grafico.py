import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tkinter import Tk, Button, filedialog, Label, messagebox, OptionMenu, StringVar
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PilImage
import os

def carregar_planilha():
    try:
        # Abre um diálogo para selecionar a planilha
        file_path = filedialog.askopenfilename(title="Selecionar Planilha", filetypes=[("Excel files", "*.xlsx;*.xls")])
        if file_path:
            processar_planilha(file_path)
        else:
            messagebox.showinfo("Informação", "Nenhum arquivo selecionado.")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao carregar a planilha: {str(e)}")

def processar_planilha(file_path):
    try:
        # Lê a planilha
        df = pd.read_excel(file_path)

        # Cria um relatório simples
        relatorio = df.describe()

        # Pergunta ao usuário qual tipo de gráfico ele quer
        tipo_grafico = var_tipo_grafico.get()

        # Gera o gráfico baseado no tipo escolhido
        plt.figure(figsize=(10, 6))
        if tipo_grafico == "Barras":
            relatorio.loc['mean'].plot(kind='bar')
        elif tipo_grafico == "Linha":
            relatorio.loc['mean'].plot(kind='line')
        elif tipo_grafico == "Histograma":
            sns.histplot(df, kde=True)
        elif tipo_grafico == "Boxplot":
            sns.boxplot(data=df)

        plt.title(f"Gráfico {tipo_grafico}")
        plt.ylabel('Valores')
        plt.xticks(rotation=45)

        # Salva o gráfico
        grafico_path = "grafico.png"
        plt.savefig(grafico_path)
        plt.close()

        # Cria um novo arquivo Excel para o relatório
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"

        # Adiciona o resumo ao relatório Excel
        for r_idx, (index, row) in enumerate(relatorio.iterrows(), 1):
            ws.cell(row=r_idx+1, column=1, value=index)
            for c_idx, value in enumerate(row, 2):
                ws.cell(row=r_idx+1, column=c_idx, value=value)

        # Adiciona o gráfico ao relatório Excel
        if os.path.exists(grafico_path):
            img = PilImage.open(grafico_path)
            img = img.resize((400, 300))  # Redimensiona a imagem para caber na planilha
            img.save(grafico_path)
            img_excel = OpenpyxlImage(grafico_path)
            ws.add_image(img_excel, 'H5')

        # Salva o relatório
        relatorio_path = 'relatorio.xlsx'
        wb.save(relatorio_path)

        # Exibe mensagem de sucesso
        messagebox.showinfo("Sucesso", f"Relatório gerado com sucesso: {relatorio_path}")

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao processar a planilha: {str(e)}")

def criar_interface():
    global var_tipo_grafico

    root = Tk()
    root.title("Gerador de Relatório Excel")

    # Label
    lbl_selecione = Label(root, text="Selecione o tipo de gráfico:")
    lbl_selecione.pack(pady=10)

    # Dropdown para selecionar tipo de gráfico
    var_tipo_grafico = StringVar(root)
    var_tipo_grafico.set("Barras")  # Valor padrão
    opcoes_grafico = ["Barras", "Linha", "Histograma", "Boxplot"]
    dropdown = OptionMenu(root, var_tipo_grafico, *opcoes_grafico)
    dropdown.pack(pady=10)

    # Botão para carregar a planilha
    btn_carregar = Button(root, text="Carregar Planilha", command=carregar_planilha)
    btn_carregar.pack(pady=20)

    # Configuração da janela
    root.geometry("400x300")
    root.mainloop()

if __name__ == "__main__":
    criar_interface()
